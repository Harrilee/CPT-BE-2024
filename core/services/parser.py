import random
from typing import List

import openpyxl

from core.services.gameModel import TypeChoiceQuestion, ThoughtChoiceQuestion, Thought, Scenario, End

import os
from django.conf import settings


def generate_random_display_id(num: int) -> List[int]:
    display_id = [i for i in range(num)]
    random.shuffle(display_id)
    return display_id


def get_scenario_list():


    wb_obj = openpyxl.load_workbook(os.path.join(settings.BASE_DIR, 'core', 'services', 'content.xlsx'))

    sheet_obj = wb_obj.active

    max_row = sheet_obj.max_row
    max_column = sheet_obj.max_column

    # Start from the last row and work upwards
    for i in range(max_row, 0, -1):
        if all(sheet_obj.cell(row=i, column=j).value is None for j in range(1, max_column + 1)):
            sheet_obj.delete_rows(i)
        else:
            break  # Stop as soon as you encounter a non-empty row

    scenario_list = []

    # scenario_count =14
    # scenario_current=0

    scenario_index = []

    for r in range(2, sheet_obj.max_row + 1):
        if sheet_obj.cell(row=r, column=1).value is not None:
            scenario_index.append(r)
    scenario_index.append(sheet_obj.max_row + 1)

    ##
    SCENARIO_COUNT = len(scenario_index) - 1
    ##



    random_display_id = generate_random_display_id(SCENARIO_COUNT)


    for i in range(len(scenario_index) - 1):  # 14 scenarios
        scenario_start_sentence = sheet_obj.cell(row=scenario_index[i], column=11).value
        scenario_end_sentence = sheet_obj.cell(row=scenario_index[i], column=10).value
        scenario_content = sheet_obj.cell(row=scenario_index[i], column=1).value
        patient_name = sheet_obj.cell(row=scenario_index[i], column=14).value

        thoughts = []
        thoughts_index = []

        for r in range(scenario_index[i], scenario_index[i + 1]):
            if sheet_obj.cell(row=r, column=2).value is not None:
                thoughts_index.append(r)
        thoughts_index.append(scenario_index[i + 1])

        for j in range(len(thoughts_index) - 1):  # j-1 thoughts

            type_choice_questions = []

            思维偏差 = sheet_obj.cell(row=thoughts_index[j], column=2).value
            alternative_thought = sheet_obj.cell(row=thoughts_index[j], column=8).value

            思维方式 = sheet_obj.cell(row=thoughts_index[j], column=3).value
            思维方式错误选项 = sheet_obj.cell(row=thoughts_index[j], column=4).value
            其他解读 = sheet_obj.cell(row=thoughts_index[j], column=8).value

            count = 0
            choices = []
            false_dict = {}
            for r in range(thoughts_index[j], thoughts_index[j + 1]):
                choices.append(sheet_obj.cell(row=r, column=6).value)
                false_dict[sheet_obj.cell(row=r, column=6).value] = sheet_obj.cell(row=r, column=9).value

            # print(choices)
            for r in range(thoughts_index[j], thoughts_index[j + 1]):
                问题类型 = sheet_obj.cell(row=r, column=5).value
                correct_choice = sheet_obj.cell(row=r, column=6).value
                true_reply = sheet_obj.cell(row=r, column=7).value

                false_reply = {}
                for k, v in false_dict.items():
                    if k != correct_choice:
                        false_reply[k] = v


                type_cq = TypeChoiceQuestion(i, random_display_id[i], 问题类型, choices, {correct_choice: true_reply},
                                             false_reply, count + 1)
                count += 1
                type_choice_questions.append(type_cq)

            thought_choice_question = ThoughtChoiceQuestion(i, random_display_id[i], 思维偏差,
                                                            [思维方式, 思维方式错误选项],
                                                            type_choice_questions[:], 其他解读)

            thought = Thought(i, random_display_id[i], 思维偏差, thought_choice_question, alternative_thought)

            thoughts.append(thought)

        scenario = Scenario(patient_name, i, random_display_id[i], scenario_start_sentence, scenario_end_sentence,
                            scenario_content, thoughts[:], next_node=End())
        scenario_list.append(scenario)
    scenario_list.sort(key=lambda x: x.display_id)

    # print('random_display_id',random_display_id)



    return scenario_list




if __name__ == '__main__':
    generate_random_display_id(14)